import openpyxl, json
from collections import defaultdict, Counter

wb = openpyxl.load_workbook("Projection_Library_Spec_v1.1.xlsx", data_only=True)
print("Sheets:", wb.sheetnames)

NOTE_PREFIXES = ("•", "Totali", "Note", "Conteggio")


def is_note_row(row_values):
    for cell in row_values:
        if cell is None:
            continue
        s = str(cell).strip()
        if s == "":
            continue
        return isinstance(cell, str) and s.startswith(NOTE_PREFIXES)
    return False


def is_empty_row(row_values):
    return not any(cell is not None and str(cell).strip() != "" for cell in row_values)


def find_header_row(ws):
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        non_empty = [c for c in row if c.value is not None and str(c.value).strip() != ""]
        if len(non_empty) >= 2:
            return row[0].row
    return 1


def count_data_rows(ws, header_row):
    n = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if is_empty_row(row) or is_note_row(row):
            continue
        n += 1
    return n


for name in wb.sheetnames:
    ws = wb[name]
    hdr = find_header_row(ws)
    data_rows = count_data_rows(ws, hdr)
    print(f"{name}: {ws.max_row} rows total, {ws.max_column} cols, header@row{hdr}, {data_rows} data rows")

ws = wb["02_voices"]
hdr_row = find_header_row(ws)
header = [cell.value for cell in ws[hdr_row]]
print(f"\n02_voices header (row {hdr_row}): {header}")


def unique_values(column_name):
    if column_name not in header:
        return None
    idx = header.index(column_name)
    values = []
    seen = set()
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if is_empty_row(row) or is_note_row(row):
            continue
        val = row[idx]
        if val is None or str(val).strip() == "":
            continue
        if val not in seen:
            seen.add(val)
            values.append(val)
    return values


for col in ("nature", "calc_phase"):
    vals = unique_values(col)
    if vals is None:
        print(f"02_voices: column '{col}' not found.")
    else:
        print(f"02_voices unique '{col}' ({len(vals)}): {vals}")


nat_idx = header.index("nature")
nat_counter = Counter()
for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
    if is_empty_row(row) or is_note_row(row):
        continue
    nat = row[nat_idx]
    if nat is None or str(nat).strip() == "":
        nat = "<missing>"
    nat_counter[nat] += 1
print(f"\n02_voices rows by 'nature' (total {sum(nat_counter.values())}):")
for nat, n in nat_counter.most_common():
    print(f"  {nat}: {n}")


def is_section_marker(row_values):
    non_empty = [c for c in row_values if c is not None and str(c).strip() != ""]
    return len(non_empty) == 1 and isinstance(non_empty[0], str) and "Sezione" in non_empty[0]


def is_notes_marker(row_values):
    non_empty = [c for c in row_values if c is not None and str(c).strip() != ""]
    return len(non_empty) == 1 and isinstance(non_empty[0], str) and non_empty[0].strip().lower().startswith("note")


def count_section_data(ws, header_row, stop_row):
    n = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_row=stop_row - 1, values_only=True):
        if is_empty_row(row) or is_note_row(row):
            continue
        n += 1
    return n


print()
ws_m = wb["01_methods"]
sections = []
for r in range(1, ws_m.max_row + 1):
    row = [c.value for c in ws_m[r]]
    if is_section_marker(row) or is_notes_marker(row):
        sections.append((r, row[0]))
sections.append((ws_m.max_row + 1, "<EOF>"))

for i in range(len(sections) - 1):
    sec_row, sec_label = sections[i]
    next_row = sections[i + 1][0]
    header_r = sec_row + 1
    while header_r < next_row:
        hdr_vals = [c.value for c in ws_m[header_r]]
        non_empty = [c for c in hdr_vals if c is not None and str(c).strip() != ""]
        if len(non_empty) >= 2:
            break
        header_r += 1
    if header_r >= next_row:
        continue
    n = count_section_data(ws_m, header_r, next_row)
    print(f"01_methods | {sec_label}: header@row{header_r}, {n} data rows")

ws_k = wb["03_kpis"]
hdr_k = find_header_row(ws_k)
print(f"\n03_kpis: header@row{hdr_k}, {count_data_rows(ws_k, hdr_k)} data rows")

ws_v = wb["05_validation"]
hdr_v = find_header_row(ws_v)
header_v = [c.value for c in ws_v[hdr_v]]
sev_idx = header_v.index("severity")
sev_counter = Counter()
for row in ws_v.iter_rows(min_row=hdr_v + 1, values_only=True):
    if is_empty_row(row) or is_note_row(row):
        continue
    sev = row[sev_idx]
    if sev is None or str(sev).strip() == "":
        sev = "<missing>"
    sev_counter[sev] += 1
print(f"\n05_validation data rows by 'severity' (total {sum(sev_counter.values())}):")
for sev, n in sev_counter.most_common():
    print(f"  {sev}: {n}")
