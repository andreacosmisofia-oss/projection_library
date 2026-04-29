"""Integration tests for the M13 Excel export endpoint.

Reuses the demo loader to seed a project + snapshot end-to-end, then
exercises ``GET /api/projects/{id}/export/excel`` and inspects the
returned workbook in-memory via openpyxl.
"""

from __future__ import annotations

import io
from typing import Iterator

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.api.main import app
from backend.infrastructure.db.database import Base, get_db


def _make_factory():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(
        bind=engine, autoflush=False, autocommit=False, future=True
    )
    return engine, factory


@pytest.fixture()
def client_and_factory() -> Iterator[tuple[TestClient, sessionmaker]]:
    engine, factory = _make_factory()

    def _override():
        db = factory()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = _override
    with TestClient(app) as c:
        yield c, factory
    app.dependency_overrides.pop(get_db, None)
    engine.dispose()


def _seed_demo(client: TestClient) -> str:
    resp = client.post("/api/demo/load-sample")
    assert resp.status_code == 201, resp.text
    return resp.json()["project_id"]


def test_export_excel_returns_xlsx_with_three_sheets(client_and_factory):
    client, _ = client_and_factory
    project_id = _seed_demo(client)

    resp = client.get(f"/api/projects/{project_id}/export/excel")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    cd = resp.headers.get("content-disposition", "")
    assert "attachment" in cd
    assert f"projection_{project_id}.xlsx" in cd

    payload = resp.content
    assert len(payload) > 1000, f"workbook suspiciously small: {len(payload)} bytes"

    wb = load_workbook(filename=io.BytesIO(payload), read_only=True)
    assert wb.sheetnames == ["P&L", "SP", "CF"]

    expected_header = ["voice_id", "Y-1", "Y0", "Y1", "Y2", "Y3"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        assert rows, f"sheet '{sheet_name}' is empty (no header row)"
        assert list(rows[0]) == expected_header, (
            f"sheet '{sheet_name}' header mismatch: {rows[0]}"
        )

    pl_rows = list(wb["P&L"].iter_rows(values_only=True))
    sp_rows = list(wb["SP"].iter_rows(values_only=True))
    assert len(pl_rows) > 1, "P&L sheet has no data rows"
    assert len(sp_rows) > 1, "SP sheet has no data rows"

    pl_voice_ids = {row[0] for row in pl_rows[1:]}
    sp_voice_ids = {row[0] for row in sp_rows[1:]}
    assert any(v.startswith("pl.") for v in pl_voice_ids)
    assert any(v.startswith("bs.") for v in sp_voice_ids)
    assert not (pl_voice_ids & sp_voice_ids), (
        "voices leaked across sheets: PL/SP overlap"
    )


def test_export_excel_returns_404_when_no_snapshot(client_and_factory):
    client, _ = client_and_factory

    create = client.post(
        "/api/projects",
        json={"name": "no-run", "sector_pack": "industrial"},
    )
    assert create.status_code == 201
    project_id = create.json()["id"]

    resp = client.get(f"/api/projects/{project_id}/export/excel")
    assert resp.status_code == 404


def test_export_excel_returns_404_when_project_missing(client_and_factory):
    client, _ = client_and_factory
    resp = client.get("/api/projects/does-not-exist/export/excel")
    assert resp.status_code == 404
