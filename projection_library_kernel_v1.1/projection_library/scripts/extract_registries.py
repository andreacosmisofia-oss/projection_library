"""Extract registry YAMLs from Projection_Library_Spec_v1.1.xlsx.

M1.2 — currently extracts:
  - 02_voices    → registries/voice_registry.yaml
  - 01_methods   → registries/method_registry.yaml (Sezione A + Sezione B)
  - 03_kpis      → registries/kpi_registry.yaml
  - 05_validation → registries/validation_rules.yaml

Validates each record against the corresponding schema in data_contracts/registries/.
Records that fail validation are reported on stderr and skipped; valid records are
emitted to YAML.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import yaml
from jsonschema import Draft202012Validator

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "Projection_Library_Spec_v1.1.xlsx"
SCHEMA_DIR = ROOT / "data_contracts" / "registries"
OUT_DIR = ROOT / "registries"

# Audit M1.0 §8.2 — composite Excel values → atomic enum.
NATURE_MAP = {
    "driver": "driver",
    "driver (placeholder)": "driver_placeholder",
    "driver (slot)": "driver_slot",
    "derived": "derived",
    "derived (identity)": "derived_identity",
    "reference": "reference",
    "placeholder": "placeholder",
}

# Audit M1.0 §9 — sign normalization.
SIGN_MAP = {
    "signed_either": "neutral",
}

# Audit M1.0 §9 — recurrence normalization.
RECURRENCE_MAP = {
    "non_recurring_extraordinary": "non_recurring",
}

# Audit M1.0 §8.3 — composite calc_phase → (proxy, final).
CALC_PHASE_MAP = {
    "E1": (None, "E1"),
    "E2": (None, "E2"),
    "E3": (None, "E3"),
    "E3.1": (None, "E3_1"),
    "E4": (None, "E4"),
    "E5": (None, "E5"),
    "E6": (None, "E6"),
    "E7": (None, "E7"),
    "E7.5": (None, "E7_5"),
    "E8": (None, "E8"),
    "E1 proxy / E3 final": ("E1", "E3"),
    "E1 proxy / E3.1 final": ("E1", "E3_1"),
    "E4 proxy / E8 check": ("E4", "E8"),
    "E7.5 final (E3 proxy)": ("E3", "E7_5"),
}

VOICES_SHEET = "02_voices"
VOICES_HEADER_ROW = 4
VOICES_DATA_START = 5

# 01_methods — sezione A (metodi) e sezione B (derived rules) sullo stesso sheet.
METHODS_SHEET = "01_methods"
METHODS_A_HEADER_ROW = 4
METHODS_A_DATA_RANGE = (5, 66)   # inclusive
METHODS_B_HEADER_ROW = 69
METHODS_B_DATA_RANGE = (70, 88)  # inclusive

# 03_kpis
KPIS_SHEET = "03_kpis"
KPIS_HEADER_ROW = 4
KPIS_DATA_RANGE = (5, 88)        # inclusive (84 KPI; audit §8)
KPI_TEMPLATE_SUFFIX = " (template)"

# 05_validation
VALIDATION_SHEET = "05_validation"
VALIDATION_HEADER_ROW = 4
VALIDATION_DATA_RANGE = (5, 77)  # inclusive (73 regole; audit §8)

# Schema enum trigger_phases (fonte verità per la normalizzazione del suffisso _post).
TRIGGER_PHASES_ENUM = {
    "E0_pre", "E0_post", "E1_post", "E2_post", "E3_post", "E3_1_post",
    "E4_post", "E5_post", "E6_post", "E7_post", "E8_post", "end_of_run",
}

# Audit M1.0 §9 — action normalization (varianti libere → enum base + qualificatore in action_note).
ACTION_MAP = {
    "halt": "halt",
    "halt_execution": "halt",
    "halt o default": "halt",
    "halt se Y0; warning altrimenti": "halt",
    "use_fallback": "use_fallback",
    "use_fallback (chain) o halt": "use_fallback",
    "use_fallback o halt": "use_fallback",
    "use_fallback (= 0 + warning)": "use_fallback",
    "use_fallback (KPI = N/A)": "use_fallback",
    "log_and_continue": "log_and_continue",
    "log_and_continue (CRITICAL)": "log_and_continue",
    "log_warning": "log_warning",
    "log_info": "log_info",
    "request_user_input": "request_user_input",
}

# Audit M1.0 §9 — complexity normalization.
COMPLEXITY_MAP = {
    "simple": "low",
    "complex": "high",
}

# Audit M1.0 §9 — family normalization (UPPERCASE → lowercase, con caso speciale BUILDUP).
FAMILY_SPECIAL_MAP = {
    "BUILDUP": "build_up",
}


def _normalize_family(value):
    if value is None or not isinstance(value, str):
        return value
    if value in FAMILY_SPECIAL_MAP:
        return FAMILY_SPECIAL_MAP[value]
    return value.lower()

# Audit M1.0 §10 — fase derived_rules → (fase_proxy, fase_final).
# Casi compositi da §10 + valori atomici secondo pattern §8.3.
FASE_MAP = {
    "E2": (None, "E2"),
    "E3.1": (None, "E3_1"),
    "E4": (None, "E4"),
    "E5": (None, "E5"),
    "E7": (None, "E7"),
    "E8": (None, "E8"),
    "E4 (proxy) → E8 (check)": ("E4", "E8"),
    "varie fasi": (None, "varies"),
}

FOOTER_PREFIXES = ("Totali", "•", "Note")


def _clean(value):
    """Normalize cell values: strip strings, treat empty / em-dash as None."""
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        if v == "" or v == "—":
            return None
        return v
    return value


def _is_footer(value: str | None) -> bool:
    if value is None:
        return False
    return any(value.startswith(p) for p in FOOTER_PREFIXES)


def extract_voices(ws) -> list[dict]:
    headers = [c.value for c in ws[VOICES_HEADER_ROW]]
    expected = [
        "voice_id", "statement", "section", "nature", "sign", "default_method",
        "calc_phase", "denominator/flow voice", "recurrence", "note",
    ]
    if headers[: len(expected)] != expected:
        raise RuntimeError(f"Unexpected headers in {VOICES_SHEET}: {headers}")

    records: list[dict] = []
    for row in ws.iter_rows(min_row=VOICES_DATA_START, values_only=True):
        voice_id = _clean(row[0])
        if voice_id is None:
            break  # first empty row → stop
        if _is_footer(voice_id):
            continue

        statement = _clean(row[1])
        section = _clean(row[2])
        nature_raw = _clean(row[3])
        sign = _clean(row[4])
        default_method = _clean(row[5])
        calc_phase_raw = _clean(row[6])
        denom = _clean(row[7])
        recurrence = _clean(row[8])
        note = _clean(row[9])

        nature = NATURE_MAP.get(nature_raw, nature_raw)
        sign = SIGN_MAP.get(sign, sign)
        recurrence = RECURRENCE_MAP.get(recurrence, recurrence)
        if calc_phase_raw is None:
            calc_phase_proxy, calc_phase_final = None, None
        elif calc_phase_raw in CALC_PHASE_MAP:
            calc_phase_proxy, calc_phase_final = CALC_PHASE_MAP[calc_phase_raw]
        else:
            # Unknown composite → leave raw in final to surface the issue via schema validation.
            calc_phase_proxy, calc_phase_final = None, calc_phase_raw

        records.append({
            "voice_id": voice_id,
            "statement": statement,
            "section": section,
            "nature": nature,
            "sign": sign,
            "default_method": default_method,
            "calc_phase_proxy": calc_phase_proxy,
            "calc_phase_final": calc_phase_final,
            "denominator_flow_voice": denom,
            "recurrence": recurrence,
            "note": note,
        })
    return records


def _validate(records: list[dict], item_schema: dict) -> tuple[list[dict], list[tuple[dict, str]]]:
    validator = Draft202012Validator(item_schema)
    valid: list[dict] = []
    invalid: list[tuple[dict, str]] = []
    for rec in records:
        errors = sorted(validator.iter_errors(rec), key=lambda e: e.path)
        if errors:
            msg = "; ".join(f"{list(e.path) or '<root>'}: {e.message}" for e in errors)
            invalid.append((rec, msg))
        else:
            valid.append(rec)
    return valid, invalid


def _report_failures(label: str, id_field: str, invalid: list[tuple[dict, str]]) -> None:
    if not invalid:
        return
    print(f"{label} — VALIDATION FAILURES — {len(invalid)} record(s):", file=sys.stderr)
    for rec, msg in invalid:
        print(f"  - {id_field}={rec.get(id_field)!r}: {msg}", file=sys.stderr)
        print(f"    record: {rec}", file=sys.stderr)


def extract_methods_section_a(ws) -> list[dict]:
    headers = [c.value for c in ws[METHODS_A_HEADER_ROW]]
    expected = [
        "method_id", "tech_code", "family", "name_it", "formula (sintesi)",
        "applicable_voices (esempi)", "complexity", "default_kpi (esempio)",
        "fallback", "note",
    ]
    if headers[: len(expected)] != expected:
        raise RuntimeError(f"Unexpected headers (sez. A) in {METHODS_SHEET}: {headers}")

    start, end = METHODS_A_DATA_RANGE
    records: list[dict] = []
    for row in ws.iter_rows(min_row=start, max_row=end, values_only=True):
        method_id = _clean(row[0])
        if method_id is None:
            continue
        if _is_footer(method_id):
            continue

        tech_code = _clean(row[1])
        family = _clean(row[2])
        name_it = _clean(row[3])
        formula_sintesi = _clean(row[4])
        applicable_raw = _clean(row[5])
        complexity_raw = _clean(row[6])
        default_kpi = _clean(row[7])
        fallback = _clean(row[8])
        note = _clean(row[9])

        if applicable_raw is None:
            applicable = None
        else:
            applicable = [item.strip() for item in applicable_raw.split(",") if item.strip()]

        complexity = COMPLEXITY_MAP.get(complexity_raw, complexity_raw)
        family = _normalize_family(family)

        records.append({
            "method_id": method_id,
            "tech_code": tech_code,
            "family": family,
            "name_it": name_it,
            "formula_sintesi": formula_sintesi,
            "applicable_voices_examples": applicable,
            "complexity": complexity,
            "default_kpi_example": default_kpi,
            "fallback": fallback,
            "note": note,
        })
    return records


def extract_methods_section_b(ws) -> list[dict]:
    headers = [c.value for c in ws[METHODS_B_HEADER_ROW]]
    expected = [
        "derived_rule_id", "tech_code", "name_it", "formula", "applicato a",
        "fase", "note",
    ]
    if headers[: len(expected)] != expected:
        raise RuntimeError(f"Unexpected headers (sez. B) in {METHODS_SHEET}: {headers}")

    start, end = METHODS_B_DATA_RANGE
    records: list[dict] = []
    for row in ws.iter_rows(min_row=start, max_row=end, values_only=True):
        rule_id = _clean(row[0])
        if rule_id is None:
            continue
        if _is_footer(rule_id):
            continue

        tech_code = _clean(row[1])
        name_it = _clean(row[2])
        formula = _clean(row[3])
        applicato_a = _clean(row[4])
        fase_raw = _clean(row[5])
        note = _clean(row[6])

        if fase_raw is None:
            fase_proxy, fase_final = None, None
        elif fase_raw in FASE_MAP:
            fase_proxy, fase_final = FASE_MAP[fase_raw]
        else:
            # Unknown → leave raw in fase_final to surface the issue via schema validation.
            fase_proxy, fase_final = None, fase_raw

        records.append({
            "derived_rule_id": rule_id,
            "tech_code": tech_code,
            "name_it": name_it,
            "formula": formula,
            "applicato_a": applicato_a,
            "fase_proxy": fase_proxy,
            "fase_final": fase_final,
            "note": note,
        })
    return records


def extract_kpis(ws) -> list[dict]:
    headers = [c.value for c in ws[KPIS_HEADER_ROW]]
    expected = [
        "kpi_id", "famiglia", "numerator", "denominator", "unit",
        "sign_handling", "default_aggregation", "used_in_methods (esempi)",
    ]
    if headers[: len(expected)] != expected:
        raise RuntimeError(f"Unexpected headers in {KPIS_SHEET}: {headers}")

    start, end = KPIS_DATA_RANGE
    records: list[dict] = []
    for row in ws.iter_rows(min_row=start, max_row=end, values_only=True):
        kpi_id = _clean(row[0])
        if kpi_id is None:
            continue
        if _is_footer(kpi_id):
            continue

        famiglia = _clean(row[1])
        numerator = _clean(row[2])
        denominator = _clean(row[3])
        unit = _clean(row[4])
        sign_handling = _clean(row[5])
        default_agg = _clean(row[6])
        used_in_raw = _clean(row[7])

        if used_in_raw is None:
            used_in = None
        else:
            used_in = [item.strip() for item in used_in_raw.split(",") if item.strip()]

        is_template = kpi_id.endswith(KPI_TEMPLATE_SUFFIX)

        records.append({
            "kpi_id": kpi_id,
            "famiglia": famiglia,
            "numerator": numerator,
            "denominator": denominator,
            "unit": unit,
            "sign_handling": sign_handling,
            "default_aggregation": default_agg,
            "used_in_methods_examples": used_in,
            "is_template": is_template,
        })
    return records


def _normalize_trigger_phase_token(token: str) -> str:
    """Garantisce che ogni token ricada nell'enum schema trigger_phases.

    - strip whitespace
    - normalizza 'E3.1' → 'E3_1' prima del controllo enum
    - se il token (eventualmente già con suffisso _post / _pre / end_of_run) è
      nell'enum schema, lo restituisce invariato
    - altrimenti aggiunge il suffisso _post (default per fasi engine).
    """
    t = token.strip().replace("E3.1", "E3_1")
    if t in TRIGGER_PHASES_ENUM:
        return t
    return t + "_post"


def _normalize_action(raw: str) -> tuple[str, str | None]:
    """Restituisce (action, action_note). Vedi ACTION_MAP + audit §9."""
    if raw in ACTION_MAP:
        action = ACTION_MAP[raw]
        action_note = None if action == raw else raw
        return action, action_note
    return "not_applicable", raw


def extract_validation_rules(ws) -> list[dict]:
    headers = [c.value for c in ws[VALIDATION_HEADER_ROW]]
    expected = [
        "rule_id", "categoria", "severity", "trigger phase", "scope",
        "regola (sintesi formula)", "action", "note",
    ]
    if headers[: len(expected)] != expected:
        raise RuntimeError(f"Unexpected headers in {VALIDATION_SHEET}: {headers}")

    start, end = VALIDATION_DATA_RANGE
    records: list[dict] = []
    for row in ws.iter_rows(min_row=start, max_row=end, values_only=True):
        rule_id = _clean(row[0])
        if rule_id is None:
            continue
        if _is_footer(rule_id):
            continue

        categoria = _clean(row[1])
        severity = _clean(row[2])
        trigger_raw = _clean(row[3])
        scope_raw = _clean(row[4])
        regola_sintesi = _clean(row[5])
        action_raw = _clean(row[6])
        note = _clean(row[7])

        if trigger_raw is None:
            trigger_phases: list[str] = []
        else:
            trigger_phases = [_normalize_trigger_phase_token(t) for t in trigger_raw.split("/") if t.strip()]

        if scope_raw is None:
            scope: list[str] = []
        else:
            scope = [s.strip() for s in scope_raw.split(",") if s.strip()]

        action, action_note = _normalize_action(action_raw) if action_raw is not None else ("not_applicable", None)

        records.append({
            "rule_id": rule_id,
            "categoria": categoria,
            "severity": severity,
            "trigger_phases": trigger_phases,
            "scope": scope,
            "regola_sintesi": regola_sintesi,
            "action": action,
            "action_note": action_note,
            "note": note,
        })
    return records


def main() -> int:
    voice_schema = json.loads((SCHEMA_DIR / "voice_registry.schema.json").read_text(encoding="utf-8"))
    method_schema = json.loads((SCHEMA_DIR / "method_registry.schema.json").read_text(encoding="utf-8"))
    kpi_schema = json.loads((SCHEMA_DIR / "kpi_registry.schema.json").read_text(encoding="utf-8"))
    validation_schema = json.loads((SCHEMA_DIR / "validation_rule.schema.json").read_text(encoding="utf-8"))

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # --- 02_voices ---
    voices = extract_voices(wb[VOICES_SHEET])
    voices_valid, voices_invalid = _validate(voices, voice_schema["$defs"]["VoiceEntry"])
    _report_failures("02_voices", "voice_id", voices_invalid)

    voices_out = OUT_DIR / "voice_registry.yaml"
    with voices_out.open("w", encoding="utf-8") as f:
        yaml.safe_dump({"voices": voices_valid}, f, sort_keys=False, allow_unicode=True, width=120)
    print(f"02_voices: {len(voices_valid)}/{len(voices)} record validi → {voices_out.relative_to(ROOT)}")

    # --- 01_methods ---
    ws_methods = wb[METHODS_SHEET]
    methods = extract_methods_section_a(ws_methods)
    methods_valid, methods_invalid = _validate(methods, method_schema["$defs"]["MethodEntry"])
    _report_failures("01_methods sez. A", "method_id", methods_invalid)

    derived = extract_methods_section_b(ws_methods)
    derived_valid, derived_invalid = _validate(derived, method_schema["$defs"]["DerivedRuleEntry"])
    _report_failures("01_methods sez. B", "derived_rule_id", derived_invalid)

    methods_out = OUT_DIR / "method_registry.yaml"
    with methods_out.open("w", encoding="utf-8") as f:
        yaml.safe_dump(
            {"methods": methods_valid, "derived_rules": derived_valid},
            f,
            sort_keys=False,
            allow_unicode=True,
            width=120,
        )
    print(f"01_methods sez. A (methods):       {len(methods_valid)}/{len(methods)} record validi → {methods_out.relative_to(ROOT)}")
    print(f"01_methods sez. B (derived_rules): {len(derived_valid)}/{len(derived)} record validi → {methods_out.relative_to(ROOT)}")

    # --- 03_kpis ---
    kpis = extract_kpis(wb[KPIS_SHEET])
    kpis_valid, kpis_invalid = _validate(kpis, kpi_schema["$defs"]["KpiEntry"])
    _report_failures("03_kpis", "kpi_id", kpis_invalid)

    kpis_out = OUT_DIR / "kpi_registry.yaml"
    with kpis_out.open("w", encoding="utf-8") as f:
        yaml.safe_dump({"kpis": kpis_valid}, f, sort_keys=False, allow_unicode=True, width=120)
    template_count = sum(1 for r in kpis_valid if r["is_template"])
    print(
        f"03_kpis: {len(kpis_valid)}/{len(kpis)} record validi "
        f"(di cui {template_count} template) → {kpis_out.relative_to(ROOT)}"
    )

    # --- 05_validation ---
    rules = extract_validation_rules(wb[VALIDATION_SHEET])
    rules_valid, rules_invalid = _validate(rules, validation_schema["$defs"]["ValidationRuleEntry"])
    _report_failures("05_validation", "rule_id", rules_invalid)

    rules_out = OUT_DIR / "validation_rules.yaml"
    with rules_out.open("w", encoding="utf-8") as f:
        yaml.safe_dump(
            {"validation_rules": rules_valid},
            f,
            sort_keys=False,
            allow_unicode=True,
            width=120,
        )
    print(f"05_validation: {len(rules_valid)}/{len(rules)} record validi → {rules_out.relative_to(ROOT)}")

    failed = bool(voices_invalid or methods_invalid or derived_invalid or kpis_invalid or rules_invalid)
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
