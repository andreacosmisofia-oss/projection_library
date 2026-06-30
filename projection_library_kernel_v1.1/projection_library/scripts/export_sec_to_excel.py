"""Export SEC EDGAR 10-K JSON data to Excel workbooks.

Produces:
  sec_edgar_10k/xlsx/{TICKER}_10K.xlsx   — one workbook per company
  sec_edgar_10k/xlsx/ALL_COMPANIES.xlsx  — consolidated multi-company view

Each company workbook has four sheets:
  Income Statement  |  Balance Sheet  |  Cash Flow  |  Key Ratios

Run after generate_sec_synthetic.py (or download_sec_edgar_10k.py).
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR   = SCRIPT_DIR.parent / "sec_edgar_10k"
OUT_DIR    = DATA_DIR / "xlsx"
OUT_DIR.mkdir(parents=True, exist_ok=True)

TICKERS = ["CAT", "PG", "GM", "PFE", "ORCL", "T", "FCX", "FDX", "NKE", "BA"]

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_SECTION_BG  = "D6E4F0"   # light steel blue
C_SECTION_FG  = "1F3864"
C_ALT_ROW     = "EEF4FB"   # very light blue
C_WHITE       = "FFFFFF"
C_TOTAL_BG    = "2E75B6"   # mid blue
C_TOTAL_FG    = "FFFFFF"
C_RATIO_BG    = "E2EFDA"   # light green
C_NEG_FG      = "C00000"   # red for negatives
C_BORDER      = "B8CCE4"

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _font(bold=False, color=C_HEADER_FG, size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_border() -> Border:
    s = Side(style="medium", color="FFFFFF")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ---------------------------------------------------------------------------
# Cell writers
# ---------------------------------------------------------------------------

def write_title(ws, row: int, col: int, text: str, cols_merge: int = 1,
                bg=C_HEADER_BG, fg=C_HEADER_FG, size=12, bold=True) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill      = _fill(bg)
    cell.alignment = _align("center")
    cell.border    = _header_border()
    if cols_merge > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col + cols_merge - 1,
        )


def write_header(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = _font(bold=True, color=C_HEADER_FG)
    cell.fill      = _fill(C_HEADER_BG)
    cell.alignment = _align("center")
    cell.border    = _header_border()


def write_section(ws, row: int, col: int, text: str, cols: int = 1) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(name="Calibri", bold=True, color=C_SECTION_FG, size=10)
    cell.fill      = _fill(C_SECTION_BG)
    cell.alignment = _align("left")
    cell.border    = _thin_border()
    if cols > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col + cols - 1,
        )


def write_label(ws, row: int, col: int, text: str,
                alt: bool = False, bold: bool = False, indent: int = 0) -> None:
    cell = ws.cell(row=row, column=col, value=("  " * indent) + text)
    cell.font      = Font(name="Calibri", bold=bold, color="000000", size=10)
    cell.fill      = _fill(C_ALT_ROW if alt else C_WHITE)
    cell.alignment = _align("left")
    cell.border    = _thin_border()


def write_number(ws, row: int, col: int, value: float | None,
                 fmt: str = "#,##0", alt: bool = False,
                 bold: bool = False, total: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    bg  = C_TOTAL_BG if total else (C_ALT_ROW if alt else C_WHITE)
    fg  = C_TOTAL_FG if total else ("000000" if (value is None or value >= 0) else C_NEG_FG)
    cell.font           = Font(name="Calibri", bold=bold or total, color=fg, size=10)
    cell.fill           = _fill(bg)
    cell.alignment      = _align("right")
    cell.border         = _thin_border()
    cell.number_format  = fmt


def write_total(ws, row: int, col: int, value: float | None,
                fmt: str = "#,##0") -> None:
    write_number(ws, row, col, value, fmt=fmt, bold=True, total=True)


# ---------------------------------------------------------------------------
# Data extractor
# ---------------------------------------------------------------------------

def _val(gaap: dict, concept: str, idx: int = 0) -> float | None:
    """Return value for concept at period index (0=latest, 1=prior, 2=two-ago)."""
    try:
        entries = gaap[concept]["units"]
        unit_key = list(entries.keys())[0]
        return entries[unit_key][idx]["val"]
    except (KeyError, IndexError):
        return None


def _periods(gaap: dict) -> list[str]:
    """Return list of year-end period strings for the available data."""
    for concept in gaap.values():
        for unit_entries in concept["units"].values():
            return [e["end"][:4] for e in unit_entries]
    return []


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_income_statement(ws, gaap: dict, entity: str, years: list[str]) -> None:
    n = len(years)
    _set_col_widths(ws, [36] + [18] * n)

    # Title
    write_title(ws, 1, 1, f"{entity} — Income Statement (USD)", 1 + n, size=13)
    write_title(ws, 2, 1, "Fiscal Year Ended Dec/May/Jun 31", 1 + n, size=10,
                bg="2E75B6")

    # Column headers
    write_header(ws, 3, 1, "Line Item")
    for i, yr in enumerate(years):
        write_header(ws, 3, 2 + i, yr)

    # Rows  (label, concept, indent, is_section, is_total, fmt)
    rows: list[tuple] = [
        ("REVENUES", None, 0, True, False, None),
        ("Revenues / Net Sales", "Revenues", 1, False, False, "#,##0"),
        ("COST & EXPENSES", None, 0, True, False, None),
        ("Cost of Goods & Services Sold", "CostOfGoodsSoldAndServicesSold", 1, False, False, "#,##0"),
        ("Gross Profit", "GrossProfit", 0, False, True, "#,##0"),
        ("SG&A Expense", "SellingGeneralAndAdministrativeExpense", 1, False, False, "#,##0"),
        ("R&D Expense", "ResearchAndDevelopmentExpense", 1, False, False, "#,##0"),
        ("Operating Income (Loss)", "OperatingIncomeLoss", 0, False, True, "#,##0"),
        ("OTHER ITEMS", None, 0, True, False, None),
        ("Interest Expense", "InterestExpense", 1, False, False, "#,##0"),
        ("Pre-Tax Income (Loss)", "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
         1, False, False, "#,##0"),
        ("Income Tax Expense (Benefit)", "IncomeTaxExpenseBenefit", 1, False, False, "#,##0"),
        ("Net Income (Loss)", "NetIncomeLoss", 0, False, True, "#,##0"),
        ("EPS — Basic", "EarningsPerShareBasic", 1, False, False, "#,##0.00"),
        ("EPS — Diluted", "EarningsPerShareDiluted", 1, False, False, "#,##0.00"),
    ]

    r = 4
    for idx, (label, concept, indent, is_section, is_total, fmt) in enumerate(rows):
        alt = idx % 2 == 0
        if is_section:
            write_section(ws, r, 1, label, cols=1 + n)
        elif is_total:
            write_label(ws, r, 1, label, bold=True)
            for i in range(n):
                v = _val(gaap, concept, i) if concept else None
                write_total(ws, r, 2 + i, v, fmt=fmt or "#,##0")
        else:
            write_label(ws, r, 1, label, alt=alt, indent=indent)
            for i in range(n):
                v = _val(gaap, concept, i) if concept else None
                write_number(ws, r, 2 + i, v, fmt=fmt or "#,##0", alt=alt)
        r += 1

    ws.freeze_panes = "B4"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16


def build_balance_sheet(ws, gaap: dict, entity: str, years: list[str]) -> None:
    n = len(years)
    _set_col_widths(ws, [36] + [18] * n)

    write_title(ws, 1, 1, f"{entity} — Balance Sheet (USD)", 1 + n, size=13)
    write_title(ws, 2, 1, "As of Year-End", 1 + n, size=10, bg="2E75B6")
    write_header(ws, 3, 1, "Line Item")
    for i, yr in enumerate(years):
        write_header(ws, 3, 2 + i, yr)

    rows: list[tuple] = [
        ("ASSETS", None, 0, True, False),
        ("Cash & Cash Equivalents", "CashAndCashEquivalentsAtCarryingValue", 1, False, False),
        ("Accounts Receivable, Net", "AccountsReceivableNetCurrent", 1, False, False),
        ("Inventory, Net", "InventoryNet", 1, False, False),
        ("Property, Plant & Equipment, Net", "PropertyPlantAndEquipmentNet", 1, False, False),
        ("Total Assets", "Assets", 0, False, True),
        ("LIABILITIES & EQUITY", None, 0, True, False),
        ("Long-Term Debt", "LongTermDebt", 1, False, False),
        ("Retained Earnings (Deficit)", "RetainedEarningsAccumulatedDeficit", 1, False, False),
        ("Total Stockholders' Equity", "StockholdersEquity", 0, False, True),
        ("Total Liabilities & Equity", "LiabilitiesAndStockholdersEquity", 0, False, True),
    ]

    r = 4
    for idx, (label, concept, indent, is_section, is_total) in enumerate(rows):
        alt = idx % 2 == 0
        if is_section:
            write_section(ws, r, 1, label, cols=1 + n)
        elif is_total:
            write_label(ws, r, 1, label, bold=True)
            for i in range(n):
                v = _val(gaap, concept, i) if concept else None
                write_total(ws, r, 2 + i, v)
        else:
            write_label(ws, r, 1, label, alt=alt, indent=indent)
            for i in range(n):
                v = _val(gaap, concept, i) if concept else None
                write_number(ws, r, 2 + i, v, alt=alt)
        r += 1

    ws.freeze_panes = "B4"


def build_cash_flow(ws, gaap: dict, entity: str, years: list[str]) -> None:
    n = len(years)
    _set_col_widths(ws, [36] + [18] * n)

    write_title(ws, 1, 1, f"{entity} — Cash Flow Statement (USD)", 1 + n, size=13)
    write_title(ws, 2, 1, "Annual", 1 + n, size=10, bg="2E75B6")
    write_header(ws, 3, 1, "Line Item")
    for i, yr in enumerate(years):
        write_header(ws, 3, 2 + i, yr)

    rows: list[tuple] = [
        ("OPERATING ACTIVITIES", None, 0, True, False),
        ("Net Cash from Operations", "NetCashProvidedByUsedInOperatingActivities", 0, False, True),
        ("INVESTING ACTIVITIES", None, 0, True, False),
        ("Capital Expenditures (Capex)", "CapitalExpendituresIncurringObligation", 1, False, False),
        ("Net Cash from Investing", "NetCashProvidedByUsedInInvestingActivities", 0, False, True),
        ("FINANCING ACTIVITIES", None, 0, True, False),
        ("Net Cash from Financing", "NetCashProvidedByUsedInFinancingActivities", 0, False, True),
        ("NON-CASH / SUPPLEMENTAL", None, 0, True, False),
        ("Depreciation & Amortization", "DepreciationDepletionAndAmortization", 1, False, False),
    ]

    r = 4
    for idx, (label, concept, indent, is_section, is_total) in enumerate(rows):
        alt = idx % 2 == 0
        if is_section:
            write_section(ws, r, 1, label, cols=1 + n)
        elif is_total:
            write_label(ws, r, 1, label, bold=True)
            for i in range(n):
                v = _val(gaap, concept, i) if concept else None
                write_total(ws, r, 2 + i, v)
        else:
            write_label(ws, r, 1, label, alt=alt, indent=indent)
            for i in range(n):
                v = _val(gaap, concept, i) if concept else None
                write_number(ws, r, 2 + i, v, alt=alt)
        r += 1

    ws.freeze_panes = "B4"


def _safe_div(a: float | None, b: float | None) -> float | None:
    if a is None or b is None or b == 0:
        return None
    return a / b


def build_ratios(ws, gaap: dict, entity: str, years: list[str]) -> None:
    n = len(years)
    _set_col_widths(ws, [36] + [18] * n)

    write_title(ws, 1, 1, f"{entity} — Key Financial Ratios", 1 + n, size=13)
    write_title(ws, 2, 1, "Derived from US-GAAP 10-K data", 1 + n, size=10, bg="2E75B6")
    write_header(ws, 3, 1, "Ratio")
    for i, yr in enumerate(years):
        write_header(ws, 3, 2 + i, yr)

    def ratio_row(ws, r, label, values, fmt="#,##0.0%", alt=False):
        write_label(ws, r, 1, label, alt=alt,
                    bold=label.startswith("  ") is False)
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=2 + i, value=v)
            cell.number_format = fmt
            cell.font      = Font(name="Calibri", size=10,
                                  color=("000000" if v is None or v >= 0 else C_NEG_FG))
            cell.fill      = _fill(C_RATIO_BG if not alt else C_ALT_ROW)
            cell.alignment = _align("right")
            cell.border    = _thin_border()

    def ratio_row_x(ws, r, label, values, fmt="#,##0.0\"x\"", alt=False):
        write_label(ws, r, 1, label, alt=alt)
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=2 + i, value=v)
            cell.number_format = fmt
            cell.font      = Font(name="Calibri", size=10,
                                  color=("000000" if v is None or v >= 0 else C_NEG_FG))
            cell.fill      = _fill(C_RATIO_BG if not alt else C_ALT_ROW)
            cell.alignment = _align("right")
            cell.border    = _thin_border()

    r = 4

    # --- Profitability ---
    write_section(ws, r, 1, "PROFITABILITY", cols=1 + n); r += 1

    gross_margins = [_safe_div(_val(gaap, "GrossProfit", i),
                                _val(gaap, "Revenues", i)) for i in range(n)]
    ratio_row(ws, r, "Gross Margin", gross_margins); r += 1

    op_margins = [_safe_div(_val(gaap, "OperatingIncomeLoss", i),
                             _val(gaap, "Revenues", i)) for i in range(n)]
    ratio_row(ws, r, "Operating Margin", op_margins, alt=True); r += 1

    net_margins = [_safe_div(_val(gaap, "NetIncomeLoss", i),
                              _val(gaap, "Revenues", i)) for i in range(n)]
    ratio_row(ws, r, "Net Profit Margin", net_margins); r += 1

    roa = [_safe_div(_val(gaap, "NetIncomeLoss", i),
                     _val(gaap, "Assets", i)) for i in range(n)]
    ratio_row(ws, r, "Return on Assets (ROA)", roa, alt=True); r += 1

    roe = [_safe_div(_val(gaap, "NetIncomeLoss", i),
                     _val(gaap, "StockholdersEquity", i)) for i in range(n)]
    ratio_row(ws, r, "Return on Equity (ROE)", roe); r += 1

    # --- Leverage ---
    write_section(ws, r, 1, "LEVERAGE & SOLVENCY", cols=1 + n); r += 1

    debt_eq = [_safe_div(_val(gaap, "LongTermDebt", i),
                          _val(gaap, "StockholdersEquity", i)) for i in range(n)]
    ratio_row_x(ws, r, "Debt / Equity", debt_eq, alt=True); r += 1

    debt_assets = [_safe_div(_val(gaap, "LongTermDebt", i),
                              _val(gaap, "Assets", i)) for i in range(n)]
    ratio_row(ws, r, "Debt / Assets", debt_assets); r += 1

    # --- Efficiency ---
    write_section(ws, r, 1, "EFFICIENCY", cols=1 + n); r += 1

    asset_turn = [_safe_div(_val(gaap, "Revenues", i),
                             _val(gaap, "Assets", i)) for i in range(n)]
    ratio_row_x(ws, r, "Asset Turnover", asset_turn, alt=True); r += 1

    # --- Cash Flow ---
    write_section(ws, r, 1, "CASH FLOW QUALITY", cols=1 + n); r += 1

    ocf_ni = [_safe_div(_val(gaap, "NetCashProvidedByUsedInOperatingActivities", i),
                         _val(gaap, "NetIncomeLoss", i)) for i in range(n)]
    ratio_row_x(ws, r, "OCF / Net Income", ocf_ni); r += 1

    fcf = [(_val(gaap, "NetCashProvidedByUsedInOperatingActivities", i) or 0)
           + (_val(gaap, "CapitalExpendituresIncurringObligation", i) or 0)  # capex already negative in CF or positive here
           for i in range(n)]
    # FCF = OCF - Capex (capex stored as positive in our data)
    fcf2 = [
        (_val(gaap, "NetCashProvidedByUsedInOperatingActivities", i) or 0)
        - abs(_val(gaap, "CapitalExpendituresIncurringObligation", i) or 0)
        for i in range(n)
    ]
    for i, v in enumerate(fcf2):
        cell = ws.cell(row=r, column=2 + i, value=v)
        cell.number_format = "#,##0"
        cell.font      = Font(name="Calibri", size=10,
                              color=("000000" if v >= 0 else C_NEG_FG))
        cell.fill      = _fill(C_ALT_ROW)
        cell.alignment = _align("right")
        cell.border    = _thin_border()
    write_label(ws, r, 1, "Free Cash Flow (USD)", alt=True); r += 1

    ws.freeze_panes = "B4"


# ---------------------------------------------------------------------------
# Company workbook
# ---------------------------------------------------------------------------

def build_company_workbook(ticker: str) -> Path:
    facts_path = DATA_DIR / ticker / "company_facts.json"
    if not facts_path.exists():
        print(f"  SKIP {ticker}: company_facts.json not found")
        return None

    data   = json.loads(facts_path.read_text(encoding="utf-8"))
    entity = data.get("entityName", ticker)
    gaap   = data.get("facts", {}).get("us-gaap", {})
    years  = _periods(gaap)

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    ws_is = wb.create_sheet("Income Statement")
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_cf = wb.create_sheet("Cash Flow")
    ws_ra = wb.create_sheet("Key Ratios")

    build_income_statement(ws_is, gaap, entity, years)
    build_balance_sheet(ws_bs, gaap, entity, years)
    build_cash_flow(ws_cf, gaap, entity, years)
    build_ratios(ws_ra, gaap, entity, years)

    # Tab colours
    ws_is.sheet_properties.tabColor = "1F3864"
    ws_bs.sheet_properties.tabColor = "2E75B6"
    ws_cf.sheet_properties.tabColor = "70AD47"
    ws_ra.sheet_properties.tabColor = "ED7D31"

    out = OUT_DIR / f"{ticker}_10K.xlsx"
    wb.save(out)
    kb = out.stat().st_size // 1024
    print(f"  {out.relative_to(DATA_DIR.parent)}  ({kb} KB)")
    return out


# ---------------------------------------------------------------------------
# Consolidated workbook
# ---------------------------------------------------------------------------

def build_consolidated() -> Path:
    """One sheet per financial statement, all companies side by side."""
    wb = Workbook()
    wb.remove(wb.active)

    all_data: list[tuple[str, str, dict, list[str]]] = []
    for ticker in TICKERS:
        p = DATA_DIR / ticker / "company_facts.json"
        if not p.exists():
            continue
        d     = json.loads(p.read_text(encoding="utf-8"))
        gaap  = d.get("facts", {}).get("us-gaap", {})
        yrs   = _periods(gaap)
        name  = d.get("entityName", ticker)
        all_data.append((ticker, name, gaap, yrs))

    # --- Consolidated Income Statement ---
    ws = wb.create_sheet("Income Statement")
    ws.sheet_properties.tabColor = "1F3864"

    IS_CONCEPTS = [
        ("Revenues", "Revenues"),
        ("COGS", "CostOfGoodsSoldAndServicesSold"),
        ("Gross Profit", "GrossProfit"),
        ("SG&A", "SellingGeneralAndAdministrativeExpense"),
        ("R&D", "ResearchAndDevelopmentExpense"),
        ("Operating Income", "OperatingIncomeLoss"),
        ("Interest Expense", "InterestExpense"),
        ("Pre-Tax Income", "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest"),
        ("Tax", "IncomeTaxExpenseBenefit"),
        ("Net Income", "NetIncomeLoss"),
        ("EPS Diluted", "EarningsPerShareDiluted"),
    ]

    IS_TOTALS = {"Gross Profit", "Operating Income", "Net Income", "Pre-Tax Income"}

    _write_consolidated_sheet(ws, "Income Statement — All Companies (USD)", all_data, IS_CONCEPTS, IS_TOTALS)

    # --- Consolidated Balance Sheet ---
    ws2 = wb.create_sheet("Balance Sheet")
    ws2.sheet_properties.tabColor = "2E75B6"

    BS_CONCEPTS = [
        ("Cash & Equivalents", "CashAndCashEquivalentsAtCarryingValue"),
        ("Accounts Receivable", "AccountsReceivableNetCurrent"),
        ("Inventory", "InventoryNet"),
        ("PP&E Net", "PropertyPlantAndEquipmentNet"),
        ("Total Assets", "Assets"),
        ("Long-Term Debt", "LongTermDebt"),
        ("Retained Earnings", "RetainedEarningsAccumulatedDeficit"),
        ("Stockholders' Equity", "StockholdersEquity"),
    ]
    BS_TOTALS = {"Total Assets", "Stockholders' Equity"}
    _write_consolidated_sheet(ws2, "Balance Sheet — All Companies (USD)", all_data, BS_CONCEPTS, BS_TOTALS)

    # --- Consolidated Cash Flow ---
    ws3 = wb.create_sheet("Cash Flow")
    ws3.sheet_properties.tabColor = "70AD47"

    CF_CONCEPTS = [
        ("OCF", "NetCashProvidedByUsedInOperatingActivities"),
        ("Capex", "CapitalExpendituresIncurringObligation"),
        ("ICF", "NetCashProvidedByUsedInInvestingActivities"),
        ("FCF", "NetCashProvidedByUsedInFinancingActivities"),
        ("D&A", "DepreciationDepletionAndAmortization"),
    ]
    CF_TOTALS = {"OCF", "ICF"}
    _write_consolidated_sheet(ws3, "Cash Flow — All Companies (USD)", all_data, CF_CONCEPTS, CF_TOTALS)

    # --- Consolidated Key Ratios (latest year only) ---
    ws4 = wb.create_sheet("Key Ratios (Latest)")
    ws4.sheet_properties.tabColor = "ED7D31"
    _write_consolidated_ratios(ws4, all_data)

    out = OUT_DIR / "ALL_COMPANIES.xlsx"
    wb.save(out)
    kb = out.stat().st_size // 1024
    print(f"  {out.relative_to(DATA_DIR.parent)}  ({kb} KB)  [consolidated]")
    return out


def _write_consolidated_sheet(
    ws,
    title: str,
    all_data: list[tuple[str, str, dict, list[str]]],
    concepts: list[tuple[str, str]],
    totals: set[str],
) -> None:
    """Layout: rows = line items, columns = (ticker, year) pairs."""
    # Build column headers: for each company show 3 years
    col_headers: list[tuple[str, str]] = []  # (ticker, year)
    for ticker, name, gaap, yrs in all_data:
        for yr in yrs:
            col_headers.append((ticker, yr))

    n_cols = len(col_headers)
    _set_col_widths(ws, [30] + [16] * n_cols)

    # Row 1: title
    write_title(ws, 1, 1, title, 1 + n_cols, size=13)

    # Row 2: ticker groups
    col = 2
    for ticker, name, gaap, yrs in all_data:
        write_title(ws, 2, col, f"{ticker} — {name}", len(yrs),
                    bg="2E75B6", size=10)
        col += len(yrs)

    # Row 3: years
    write_header(ws, 3, 1, "Line Item")
    for c, (tk, yr) in enumerate(col_headers):
        write_header(ws, 3, 2 + c, yr)

    # Data rows
    for r_idx, (label, concept) in enumerate(concepts):
        r = 4 + r_idx
        alt = r_idx % 2 == 0
        is_total = label in totals
        if is_total:
            write_label(ws, r, 1, label, bold=True)
        else:
            write_label(ws, r, 1, label, alt=alt)
        c = 0
        for ticker, name, gaap, yrs in all_data:
            for yr_idx in range(len(yrs)):
                v = _val(gaap, concept, yr_idx)
                fmt = "#,##0.00" if "EPS" in label else "#,##0"
                if is_total:
                    write_total(ws, r, 2 + c, v, fmt=fmt)
                else:
                    write_number(ws, r, 2 + c, v, fmt=fmt, alt=alt)
                c += 1

    ws.freeze_panes = "B4"


def _write_consolidated_ratios(ws, all_data: list[tuple[str, str, dict, list[str]]]) -> None:
    n = len(all_data)
    _set_col_widths(ws, [30] + [18] * n)

    write_title(ws, 1, 1, "Key Ratios — Latest Fiscal Year (all companies)", 1 + n, size=13)
    write_header(ws, 2, 1, "Ratio")
    for i, (ticker, name, _, yrs) in enumerate(all_data):
        write_header(ws, 2, 2 + i, f"{ticker}\n{yrs[0] if yrs else ''}")
        ws.row_dimensions[2].height = 30

    def pct_row(ws, r, label, values, alt=False):
        write_label(ws, r, 1, label, alt=alt)
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=2 + i, value=v)
            cell.number_format = "#,##0.0%"
            cell.font      = Font(name="Calibri", size=10,
                                  color="000000" if v is None or v >= 0 else C_NEG_FG)
            cell.fill      = _fill(C_RATIO_BG if not alt else C_ALT_ROW)
            cell.alignment = _align("right")
            cell.border    = _thin_border()

    def mult_row(ws, r, label, values, alt=False):
        write_label(ws, r, 1, label, alt=alt)
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=2 + i, value=v)
            cell.number_format = "#,##0.0\"x\""
            cell.font      = Font(name="Calibri", size=10,
                                  color="000000" if v is None or v >= 0 else C_NEG_FG)
            cell.fill      = _fill(C_RATIO_BG if not alt else C_ALT_ROW)
            cell.alignment = _align("right")
            cell.border    = _thin_border()

    r = 3

    write_section(ws, r, 1, "PROFITABILITY", cols=1 + n); r += 1

    pct_row(ws, r, "Gross Margin",
            [_safe_div(_val(g, "GrossProfit"), _val(g, "Revenues")) for _, _, g, _ in all_data]); r += 1

    pct_row(ws, r, "Operating Margin",
            [_safe_div(_val(g, "OperatingIncomeLoss"), _val(g, "Revenues")) for _, _, g, _ in all_data],
            alt=True); r += 1

    pct_row(ws, r, "Net Margin",
            [_safe_div(_val(g, "NetIncomeLoss"), _val(g, "Revenues")) for _, _, g, _ in all_data]); r += 1

    pct_row(ws, r, "ROA",
            [_safe_div(_val(g, "NetIncomeLoss"), _val(g, "Assets")) for _, _, g, _ in all_data],
            alt=True); r += 1

    pct_row(ws, r, "ROE",
            [_safe_div(_val(g, "NetIncomeLoss"), _val(g, "StockholdersEquity")) for _, _, g, _ in all_data]); r += 1

    write_section(ws, r, 1, "LEVERAGE", cols=1 + n); r += 1

    mult_row(ws, r, "Debt / Equity",
             [_safe_div(_val(g, "LongTermDebt"), _val(g, "StockholdersEquity")) for _, _, g, _ in all_data],
             alt=True); r += 1

    pct_row(ws, r, "Debt / Assets",
            [_safe_div(_val(g, "LongTermDebt"), _val(g, "Assets")) for _, _, g, _ in all_data]); r += 1

    write_section(ws, r, 1, "CASH FLOW", cols=1 + n); r += 1

    mult_row(ws, r, "OCF / Net Income",
             [_safe_div(_val(g, "NetCashProvidedByUsedInOperatingActivities"), _val(g, "NetIncomeLoss"))
              for _, _, g, _ in all_data],
             alt=True); r += 1

    ws.freeze_panes = "B3"


def _safe_div(a, b):
    if a is None or b is None or b == 0:
        return None
    return a / b


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print(f"Output: {OUT_DIR}\n")

    for ticker in TICKERS:
        build_company_workbook(ticker)

    print()
    build_consolidated()
    print(f"\nDone. {len(TICKERS)} company workbooks + 1 consolidated.")


if __name__ == "__main__":
    main()
