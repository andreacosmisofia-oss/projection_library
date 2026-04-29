"""Excel export — Milestone M13.

Builds a 3-sheet xlsx (P&L / SP / CF) from the project's latest
snapshot. Rows are sourced from ``snapshot_values`` (already
normalised by ``build_snapshot_values``); each ``voice_id`` is routed
to a sheet via ``registries.voices[voice_id].statement`` (PL → P&L,
BS → SP, CF → CF). Columns are fixed: ``Y-1, Y0, Y1, Y2, Y3``.
"""

from __future__ import annotations

import io
from typing import Annotated, Iterable

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from backend.infrastructure.db.database import get_db
from backend.infrastructure.db.models import Project, SnapshotValue
from backend.infrastructure.db.repositories.snapshots_repo import (
    get_latest_snapshot,
)
from backend.infrastructure.registry import Registries, get_cache

router = APIRouter(prefix="/api/projects/{project_id}/export", tags=["exports"])


_EXPORT_COLUMNS: tuple[str, ...] = ("Y-1", "Y0", "Y1", "Y2", "Y3")
_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Statement (registry) → sheet name (user-facing).
_SHEET_BY_STATEMENT: dict[str, str] = {
    "PL": "P&L",
    "BS": "SP",
    "CF": "CF",
}


def _get_project_or_404(db: Session, project_id: str) -> Project:
    project = db.get(Project, project_id)
    if project is None or project.deleted_at is not None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"project '{project_id}' not found",
        )
    return project


def _registries() -> Registries:
    return get_cache().get()


def _group_values(
    values: Iterable[SnapshotValue], registries: Registries
) -> dict[str, dict[str, dict[str, float | None]]]:
    """Return ``{sheet_name: {voice_id: {year: value}}}``.

    Voices whose ``voice_id`` is unknown to the registry, or whose
    statement is not one of PL/BS/CF, are skipped — the export is
    intended as a clean view of the standard three statements.
    """
    grouped: dict[str, dict[str, dict[str, float | None]]] = {
        sheet: {} for sheet in _SHEET_BY_STATEMENT.values()
    }
    for sv in values:
        voice = registries.voices.get(sv.voice_id)
        if voice is None:
            continue
        sheet = _SHEET_BY_STATEMENT.get(voice.get("statement", ""))
        if sheet is None:
            continue
        grouped[sheet].setdefault(sv.voice_id, {})[sv.year] = sv.value
    return grouped


def _build_workbook(
    grouped: dict[str, dict[str, dict[str, float | None]]],
) -> bytes:
    """Render the grouped data into an xlsx byte string."""
    wb = Workbook()
    # The default sheet is replaced by the first explicit one to keep
    # the workbook ordered exactly as ``_SHEET_BY_STATEMENT`` declares.
    wb.remove(wb.active)

    for sheet_name in _SHEET_BY_STATEMENT.values():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["voice_id", *_EXPORT_COLUMNS])
        rows = grouped.get(sheet_name, {})
        for voice_id in sorted(rows):
            year_map = rows[voice_id]
            ws.append(
                [voice_id]
                + [year_map.get(year) for year in _EXPORT_COLUMNS]
            )
        # Modest column widening so the file is readable on first open.
        ws.column_dimensions[get_column_letter(1)].width = 40
        for idx in range(2, 2 + len(_EXPORT_COLUMNS)):
            ws.column_dimensions[get_column_letter(idx)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/excel")
def export_excel(
    project_id: str,
    db: Annotated[Session, Depends(get_db)],
) -> StreamingResponse:
    _get_project_or_404(db, project_id)
    snapshot = get_latest_snapshot(db, project_id)
    if snapshot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=(
                f"no snapshot found for project '{project_id}' — "
                f"POST /api/projects/{project_id}/run first"
            ),
        )

    grouped = _group_values(snapshot.values, _registries())
    payload = _build_workbook(grouped)

    filename = f"projection_{project_id}.xlsx"
    return StreamingResponse(
        io.BytesIO(payload),
        media_type=_XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


__all__ = ["router"]
