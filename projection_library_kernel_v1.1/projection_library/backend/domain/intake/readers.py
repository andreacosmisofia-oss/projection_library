"""File-format readers (Excel, CSV).

Each reader returns a list of *raw sheets* — a sheet is just a name
plus the 2-D matrix of cell values. The header / row extraction step
is delegated to ``parser.py``.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from pathlib import PurePath

import openpyxl

from backend.domain.intake.exceptions import ParseError


@dataclass(frozen=True)
class RawSheet:
    name: str
    rows: list[list[object | None]]


def detect_format(filename: str) -> str:
    suffix = PurePath(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return "xlsx"
    if suffix == ".csv":
        return "csv"
    raise ParseError(
        f"unsupported file extension '{suffix}' (expected .xlsx, .xls or .csv)"
    )


def read_xlsx(file_bytes: bytes) -> list[RawSheet]:
    """Load every sheet of an Excel workbook into ``RawSheet``s."""
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as exc:  # openpyxl raises a zoo of exception types
        raise ParseError(f"unable to open Excel file: {exc}") from exc

    sheets: list[RawSheet] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[object | None]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets.append(RawSheet(name=sheet_name, rows=rows))
    wb.close()
    if not sheets:
        raise ParseError("Excel file contains no sheets")
    return sheets


def read_csv(file_bytes: bytes) -> list[RawSheet]:
    """Decode a CSV file (UTF-8 with latin-1 fallback) into a single sheet."""
    text: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ParseError("CSV file is not valid UTF-8 nor latin-1")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") > sample.count(",") else ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows: list[list[object | None]] = [list(r) for r in reader]
    return [RawSheet(name="csv", rows=rows)]


__all__ = ["RawSheet", "detect_format", "read_csv", "read_xlsx"]
